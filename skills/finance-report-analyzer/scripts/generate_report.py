#!/usr/bin/env python3
"""
Generate financial analysis reports from Excel data with text analysis.

Usage:
    python3 generate_report.py <input_xlsx> [-o format] [--company NAME] [--ticker TICKER]

Output formats: html (always generated), pdf (default), doc, md
"""

import sys
import os
import argparse
import subprocess

try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl -q")
    import openpyxl


def parse_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = {}
    header_row = None
    for row in ws.iter_rows(min_row=1, values_only=False):
        vals = [cell.value for cell in row]
        label = str(vals[0]).strip() if vals[0] else ""
        if not header_row:
            header_row = vals[1:]
        rows[label] = vals[1:]
    return rows, header_row


def identify_periods(rows):
    report_types = []
    forecast_flags = []
    # Find period row (年报/季报) and forecast row (盈利预测)
    for key in rows:
        vals = [str(v) if v else "" for v in rows[key]]
        if not report_types and ("报告期" in key or any("年报" in v for v in vals)):
            if any("年报" in v or "季报" in v or "中报" in v for v in vals):
                report_types = vals
        if not forecast_flags and any("盈利预测" in v for v in vals):
            forecast_flags = vals
    # Merge: if report_types doesn't have 盈利预测 but forecast_flags does, overlay
    if report_types and forecast_flags and not any("盈利预测" in r for r in report_types):
        for i, f in enumerate(forecast_flags):
            if "盈利预测" in f and i < len(report_types):
                report_types[i] = report_types[i] + "|盈利预测" if report_types[i] else "盈利预测"
    if not report_types:
        first_key = list(rows.keys())[0]
        return list(range(len(rows[first_key]))), ["" for _ in rows[first_key]]
    annual_idx = [i for i, r in enumerate(report_types) if r.endswith("年报") and "下半年" not in r]
    forecast_idx = [i for i, r in enumerate(report_types) if "盈利预测" in r]
    return annual_idx + forecast_idx, report_types


def get_values(rows, label, indices):
    vals = rows.get(label, [])
    return [vals[i] if i < len(vals) else None for i in indices]


def fmt(v, decimals=2):
    if v is None or v == "" or str(v) == "None":
        return "-"
    try:
        f = float(v)
        return f"{f:,.{decimals}f}" if abs(f) > 1000 else f"{f:.{decimals}f}"
    except (ValueError, TypeError):
        return str(v)


def fmt_pct(v):
    if v is None or v == "" or str(v) == "None":
        return "-"
    try:
        return f"{float(v):.1f}%"
    except (ValueError, TypeError):
        return str(v)


def to_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def find_row_label(rows, keywords):
    for label in rows:
        for kw in keywords:
            if kw in label:
                return label
    return None


def sparkline_svg(values, is_forecast_flags, width=80, height=22):
    """Generate compact inline SVG sparkline."""
    nums = [to_float(v) for v in values]
    valid = [n for n in nums if n is not None]
    if len(valid) < 2:
        return ""
    mn, mx = min(valid), max(valid)
    rng = mx - mn if mx != mn else 1
    pad = 2
    points = []
    for i, n in enumerate(nums):
        if n is None:
            continue
        x = pad + (i / (len(nums) - 1)) * (width - 2 * pad) if len(nums) > 1 else width / 2
        y = pad + (1 - (n - mn) / rng) * (height - 2 * pad)
        points.append((x, y, i))

    color = "#16a34a" if valid[-1] >= valid[0] else "#dc2626"
    actual_end = len(values) - sum(1 for f in is_forecast_flags if f)
    svg = f'<svg width="{width}" height="{height}" viewBox="0 0 {width} {height}" style="vertical-align:middle;">'

    actual_pts = [(x, y) for x, y, idx in points if idx < actual_end]
    if len(actual_pts) >= 2:
        path = "M" + " L".join(f"{x:.1f},{y:.1f}" for x, y in actual_pts)
        svg += f'<path d="{path}" fill="none" stroke="{color}" stroke-width="1.5" stroke-linecap="round"/>'

    forecast_pts = [(x, y) for x, y, idx in points if idx >= actual_end]
    if actual_pts and forecast_pts:
        bridge = [actual_pts[-1]] + forecast_pts
        path = "M" + " L".join(f"{x:.1f},{y:.1f}" for x, y in bridge)
        svg += f'<path d="{path}" fill="none" stroke="#f59e0b" stroke-width="1.5" stroke-dasharray="3,2" stroke-linecap="round"/>'

    svg += "</svg>"
    return svg


def calc_cagr(start, end, years):
    if start and end and years > 0 and start != 0:
        try:
            s, e = float(start), float(end)
            if s > 0 and e > 0:
                return (pow(e / s, 1 / years) - 1) * 100
        except:
            pass
    return None


def analyze_profitability(dates, is_forecast, revenue, op_profit, net_profit, rd, gross_margin, net_margin, roe, roa):
    """Generate text analysis for profitability section."""
    actual = [(i, d) for i, (d, f) in enumerate(zip(dates, is_forecast)) if not f]
    if len(actual) < 2:
        return ""
    li, ld = actual[-1]
    pi, pd = actual[-2]
    fi, fd = actual[0]

    lines = []
    # Revenue trend
    rv_last = to_float(revenue[li])
    rv_prev = to_float(revenue[pi])
    rv_first = to_float(revenue[fi])
    if rv_last and rv_prev:
        yoy = (rv_last / rv_prev - 1) * 100
        lines.append(f"<strong>{ld}年营收{fmt(rv_last)}亿元</strong>，同比{'增长' if yoy > 0 else '下降'}{abs(yoy):.1f}%。")
    if rv_last and rv_first and len(actual) > 2:
        cagr = calc_cagr(rv_first, rv_last, len(actual) - 1)
        if cagr:
            lines.append(f"{fd}-{ld}年营收CAGR约{cagr:.0f}%，{'展现强劲增长动能' if cagr > 20 else '保持稳健增长' if cagr > 5 else '增长趋于平缓'}。")

    # Profit analysis
    np_last = to_float(net_profit[li])
    np_prev = to_float(net_profit[pi])
    if np_last is not None:
        if np_last > 0:
            lines.append(f"归母净利润{fmt(np_last)}亿元，公司处于盈利状态。")
        else:
            lines.append(f"归母净利润{fmt(np_last)}亿元，仍处于亏损状态。")
            if np_prev and np_last > np_prev:
                improvement = (1 - abs(np_last) / abs(np_prev)) * 100 if np_prev != 0 else 0
                lines.append(f"但亏损较上年收窄{improvement:.0f}%，改善趋势明显。")

    # R&D
    rd_last = to_float(rd[li])
    if rd_last and rv_last:
        rd_ratio = rd_last / rv_last * 100
        lines.append(f"研发投入{fmt(rd_last)}亿元，占营收{rd_ratio:.1f}%，{'研发强度极高，属于典型研发驱动型企业' if rd_ratio > 30 else '研发投入力度较大' if rd_ratio > 15 else '研发投入比例适中'}。")

    # Margins
    gm = to_float(gross_margin[li])
    nm = to_float(net_margin[li])
    r = to_float(roe[li])
    if gm:
        lines.append(f"毛利率{gm:.1f}%，{'处于高毛利水平，产品竞争力强' if gm > 70 else '毛利率良好' if gm > 40 else '毛利率偏低，需关注成本控制'}。")
    if r:
        lines.append(f"ROE为{r:.1f}%，{'盈利能力优秀' if r > 15 else '回报水平合理' if r > 0 else '尚未实现正向股东回报'}。")

    # Forecast hint
    forecast = [(i, d) for i, (d, f) in enumerate(zip(dates, is_forecast)) if f]
    if forecast:
        fi2, fd2 = forecast[0]
        np_fc = to_float(net_profit[fi2])
        if np_fc is not None and np_last is not None:
            if np_fc > 0 and np_last < 0:
                lines.append(f"⭐ <strong>关键拐点：{fd2}年预计实现盈利（{fmt(np_fc)}亿元），有望结束亏损历史。</strong>")
            elif np_fc > np_last:
                lines.append(f"预计{fd2}年利润将进一步改善至{fmt(np_fc)}亿元。")

    return " ".join(lines)


def analyze_balance_sheet(dates, is_forecast, total_assets, total_liab, equity, current_assets, debt_ratio):
    actual = [(i, d) for i, (d, f) in enumerate(zip(dates, is_forecast)) if not f]
    if len(actual) < 2:
        return ""
    li, ld = actual[-1]
    pi, pd = actual[-2]
    lines = []

    ta = to_float(total_assets[li])
    tl = to_float(total_liab[li])
    eq = to_float(equity[li])
    dr = to_float(debt_ratio[li])
    ca = to_float(current_assets[li])

    if ta:
        ta_prev = to_float(total_assets[pi])
        yoy = (ta / ta_prev - 1) * 100 if ta_prev else 0
        lines.append(f"{ld}年总资产{fmt(ta)}亿元，同比{'增长' if yoy > 0 else '下降'}{abs(yoy):.1f}%。")

    if dr is not None:
        if dr < 40:
            lines.append(f"资产负债率{dr:.1f}%，财务结构稳健，杠杆水平较低。")
        elif dr < 60:
            lines.append(f"资产负债率{dr:.1f}%，杠杆处于合理区间。")
        else:
            lines.append(f"资产负债率{dr:.1f}%，<strong>杠杆偏高，需关注偿债风险</strong>。")

    if ca and tl:
        cl = to_float(total_liab[li])  # approx
        if ca and cl and cl > 0:
            ratio = ca / cl
            lines.append(f"流动资产{fmt(ca)}亿元，{'短期偿债能力充足' if ratio > 1.5 else '流动性尚可' if ratio > 1 else '流动性偏紧，需关注短期偿债压力'}。")

    if eq is not None:
        eq_prev = to_float(equity[pi])
        if eq_prev:
            chg = (eq / eq_prev - 1) * 100
            if chg < -10:
                lines.append(f"股东权益同比下降{abs(chg):.1f}%，持续亏损正在侵蚀净资产。")
            elif chg > 10:
                lines.append(f"股东权益同比增长{chg:.1f}%，资本实力增强。")

    return " ".join(lines)


def analyze_cash_flow(dates, is_forecast, ocf, icf, fcf_finance, cash_end, capex):
    actual = [(i, d) for i, (d, f) in enumerate(zip(dates, is_forecast)) if not f]
    if len(actual) < 2:
        return ""
    li, ld = actual[-1]
    pi, pd = actual[-2]
    lines = []

    o = to_float(ocf[li])
    o_prev = to_float(ocf[pi])
    ce = to_float(cash_end[li])
    cx = to_float(capex[li])

    if o is not None:
        if o > 0:
            lines.append(f"{ld}年经营活动现金流{fmt(o)}亿元，<strong>已实现正向造血</strong>。")
        else:
            lines.append(f"{ld}年经营活动现金流{fmt(o)}亿元，经营尚未产生正向现金流。")
            if o_prev and o > o_prev:
                lines.append("但较上年明显改善，现金流缺口正在收窄。")

    if o and cx:
        fcf = o - cx
        lines.append(f"自由现金流{fmt(fcf)}亿元，{'具备自我造血能力' if fcf > 0 else '仍需外部融资支撑发展'}。")

    if cx:
        lines.append(f"资本开支{fmt(cx)}亿元，{'投资扩张力度较大' if cx > 30 else '维持正常资本投入'}。")

    if ce:
        lines.append(f"期末现金余额{fmt(ce)}亿元，{'现金储备充裕' if ce > 100 else '现金储备尚可' if ce > 30 else '现金储备偏低，需关注流动性'}。")

    return " ".join(lines)


def analyze_per_share(dates, is_forecast, eps, bps, asset_turnover, employees, revenue):
    actual = [(i, d) for i, (d, f) in enumerate(zip(dates, is_forecast)) if not f]
    if len(actual) < 2:
        return ""
    li, ld = actual[-1]
    lines = []

    e = to_float(eps[li])
    b = to_float(bps[li])
    at = to_float(asset_turnover[li])
    emp = to_float(employees[li])
    rev = to_float(revenue[li])

    if e is not None:
        lines.append(f"{ld}年EPS为{fmt(e)}元，{'已实现每股盈利' if e > 0 else '每股仍处亏损状态'}。")
    if b:
        lines.append(f"每股净资产{fmt(b)}元。")
    if at:
        lines.append(f"资产周转率{at:.2f}倍，{'运营效率较高' if at > 0.5 else '资产运营效率有提升空间'}。")
    if emp and rev:
        per = rev * 10000 / emp
        lines.append(f"员工{int(emp)}人，人均营收约{per:.0f}万元。")

    return " ".join(lines)


def generate_industry_analysis(company, revenue, net_profit, gross_margin, rd, dates, is_forecast):
    """Generate industry analysis section."""
    actual = [(i, d) for i, (d, f) in enumerate(zip(dates, is_forecast)) if not f]
    if not actual:
        return ""
    li, ld = actual[-1]
    gm = to_float(gross_margin[li])
    rv = to_float(revenue[li])
    rd_v = to_float(rd[li])

    lines = []
    lines.append(f"<h3>行业定位</h3>")
    if gm and gm > 80:
        lines.append(f"<p>毛利率高达{gm:.1f}%，处于行业顶尖水平，体现出强大的产品定价权和技术壁垒。在创新药行业中，80%以上的毛利率通常意味着拥有专利保护的核心产品和较强的议价能力。</p>")
    elif gm and gm > 50:
        lines.append(f"<p>毛利率{gm:.1f}%，在行业中处于中上水平。</p>")

    if rd_v and rv:
        ratio = rd_v / rv * 100
        lines.append(f"<h3>研发竞争力</h3>")
        if ratio > 30:
            lines.append(f"<p>研发营收比{ratio:.0f}%，远超行业平均水平（通常15-25%）。高研发投入是构建长期竞争护城河的关键，但同时也是短期利润承压的主要原因。随着核心产品进入商业化成熟期，研发费用率有望逐步下降，推动利润释放。</p>")
        else:
            lines.append(f"<p>研发营收比{ratio:.0f}%，研发投入处于{'较高' if ratio > 15 else '适中'}水平。</p>")

    lines.append(f"<h3>竞争格局</h3>")
    lines.append(f"<p>需关注同行业竞品的市场份额变化、管线进展及定价策略。行业整体趋势包括：集采降价压力、创新药出海加速、生物类似药竞争等。建议结合行业报告进一步对标分析。</p>")

    return "\n".join(lines)


def generate_risk_analysis(dates, is_forecast, net_profit, debt_ratio, ocf, cash_end, revenue, gross_margin):
    """Generate risk analysis section."""
    actual = [(i, d) for i, (d, f) in enumerate(zip(dates, is_forecast)) if not f]
    if not actual:
        return ""
    li, ld = actual[-1]

    risks = []
    opportunities = []

    np = to_float(net_profit[li])
    dr = to_float(debt_ratio[li])
    o = to_float(ocf[li])
    ce = to_float(cash_end[li])
    rv = to_float(revenue[li])
    gm = to_float(gross_margin[li])

    # Risks
    if np is not None and np < 0:
        risks.append(("持续亏损风险", "公司尚未实现年度盈利，若核心产品销售不及预期或研发支出持续高企，亏损周期可能延长，影响估值和融资能力。"))
    if dr and dr > 50:
        risks.append(("财务杠杆风险", f"资产负债率{dr:.1f}%偏高，若经营环境恶化可能面临偿债压力。"))
    if o is not None and o < 0:
        risks.append(("现金流风险", "经营现金流为负，依赖外部融资维持运营。需关注融资渠道和条件的变化。"))

    risks.append(("行业政策风险", "医药行业受集采降价、医保谈判等政策影响大，核心产品价格可能承压。"))
    risks.append(("研发失败风险", "在研管线存在临床试验失败或审批延迟的不确定性，可能影响未来增长预期。"))
    risks.append(("竞争加剧风险", "同靶点竞品不断涌现，市场份额和定价策略面临挑战。"))
    risks.append(("地缘政治风险", "海外业务受国际关系、贸易政策等影响，存在市场准入不确定性。"))

    # Opportunities
    if rv and to_float(revenue[actual[-2][0] if len(actual) > 1 else 0]):
        rv_prev = to_float(revenue[actual[-2][0]])
        if rv_prev and (rv / rv_prev - 1) > 0.2:
            opportunities.append(("高速增长", f"营收保持{(rv/rv_prev-1)*100:.0f}%的高增长，商业化进展顺利。"))
    if gm and gm > 80:
        opportunities.append(("高毛利优势", f"毛利率{gm:.1f}%，产品竞争力强，一旦费用率下降将快速释放利润。"))

    # Check forecast for profit turnaround
    forecast = [(i, d) for i, (d, f) in enumerate(zip(dates, is_forecast)) if f]
    if forecast and np is not None and np < 0:
        np_fc = to_float(net_profit[forecast[0][0]])
        if np_fc and np_fc > 0:
            opportunities.append(("盈利拐点", f"预计{forecast[0][1]}年实现盈利，将从根本上改变公司估值逻辑。"))

    if ce and ce > 50:
        opportunities.append(("现金储备充裕", f"账面现金{fmt(ce)}亿元，为研发和扩张提供充足弹药。"))

    opportunities.append(("全球化布局", "海外市场渗透率提升空间大，国际化战略持续推进。"))

    html = '<div class="two-col"><div><h3>🔴 风险因素</h3><ul class="risk-list">'
    for title, desc in risks:
        html += f'<li><strong>{title}：</strong>{desc}</li>'
    html += '</ul></div><div><h3>🟢 积极因素</h3><ul class="risk-list">'
    for title, desc in opportunities:
        html += f'<li><strong>{title}：</strong>{desc}</li>'
    html += '</ul></div></div>'
    return html


def generate_html(rows, header_row, indices, report_types, company="", ticker=""):
    dates = []
    for i in indices:
        h = str(header_row[i]) if i < len(header_row) and header_row[i] else ""
        dates.append(h[:4] if len(h) >= 4 else h)

    is_forecast = ["盈利预测" in report_types[i] if i < len(report_types) else False for i in indices]

    def get(keywords):
        label = find_row_label(rows, keywords)
        return get_values(rows, label, indices) if label else [None] * len(indices)

    revenue = get(["营业总收入", "营业收入"])
    op_profit = get(["营业利润"])
    net_profit_parent = get(["归属母公司股东的净利润", "归母净利润"])
    net_profit = get(["净利润"])
    rd = get(["研发支出", "研发费用"])
    ebitda = get(["EBITDA"])
    gross_margin = get(["销售毛利率", "毛利率"])
    net_margin = get(["销售净利率", "净利率"])
    roe = get(["ROE(摊薄)", "ROE"])
    roa = get(["ROA"])
    total_assets = get(["资产总计", "总资产"])
    total_liab = get(["负债合计", "总负债"])
    equity = get(["归属母公司股东的权益", "股东权益"])
    current_assets = get(["流动资产"])
    debt_ratio = get(["资产负债率"])
    ocf = get(["经营活动现金净流量", "经营活动现金流"])
    icf = get(["投资活动现金净流量"])
    fcf_finance = get(["筹资活动现金净流量"])
    cash_end = get(["期末现金余额"])
    capex = get(["购建固定无形长期资产支付的现金", "资本开支"])
    eps = get(["EPS(基本)", "EPS"])
    bps = get(["每股净资产", "BPS"])
    employees = get(["员工总数"])
    asset_turnover = get(["资产周转率"])

    fcf = []
    for o, c in zip(ocf, capex):
        o_f, c_f = to_float(o), to_float(c)
        fcf.append(o_f - c_f if o_f is not None and c_f is not None else None)

    title = f"{company} [{ticker}]" if ticker else company or "财务分析报告"

    def table_header():
        h = "<tr><th>指标</th>"
        for i, y in enumerate(dates):
            cls = ' class="forecast"' if is_forecast[i] else ""
            suffix = "E" if is_forecast[i] else ""
            h += f"<th{cls}>{y}{suffix}</th>"
        h += "<th>趋势</th></tr>"
        return h

    def table_row(label, values, is_pct=False):
        r = f"<tr><td>{label}</td>"
        for i, v in enumerate(values):
            fc = is_forecast[i]
            cls = ' class="forecast"' if fc else ""
            forecast_mark = ' <span class="forecast-dot" title="预测值">⟡</span>' if fc else ""
            val_cls_open = '<span class="forecast-val">' if fc else ""
            val_cls_close = "</span>" if fc else ""
            if is_pct:
                r += f'<td{cls}>{val_cls_open}{fmt_pct(v)}{val_cls_close}{forecast_mark}</td>'
            else:
                fv = to_float(v)
                color = ""
                if fv is not None and not fc:
                    color = ' class="positive"' if fv > 0 else ' class="negative"' if fv < 0 else ""
                elif fv is not None and fc:
                    color = ' class="positive forecast-val"' if fv > 0 else ' class="negative forecast-val"' if fv < 0 else ' class="forecast-val"'
                r += f'<td{cls}><span{color}>{fmt(v)}</span>{forecast_mark}</td>'
        spark = sparkline_svg(values, is_forecast)
        r += f"<td>{spark}</td></tr>"
        return r

    # Summary uses only actual values
    actual_indices = [i for i, f in enumerate(is_forecast) if not f]
    latest = actual_indices[-1] if actual_indices else 0
    prev = actual_indices[-2] if len(actual_indices) > 1 else 0

    rv_last = to_float(revenue[latest])
    rv_prev = to_float(revenue[prev])
    rv_yoy = f"{(rv_last/rv_prev-1)*100:.1f}%" if rv_last and rv_prev else "-"

    # Text analyses
    profit_text = analyze_profitability(dates, is_forecast, revenue, op_profit, net_profit_parent, rd, gross_margin, net_margin, roe, roa)
    bs_text = analyze_balance_sheet(dates, is_forecast, total_assets, total_liab, equity, current_assets, debt_ratio)
    cf_text = analyze_cash_flow(dates, is_forecast, ocf, icf, fcf_finance, cash_end, capex)
    ps_text = analyze_per_share(dates, is_forecast, eps, bps, asset_turnover, employees, revenue)
    industry_html = generate_industry_analysis(company, revenue, net_profit_parent, gross_margin, rd, dates, is_forecast)
    risk_html = generate_risk_analysis(dates, is_forecast, net_profit_parent, debt_ratio, ocf, cash_end, revenue, gross_margin)

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{title} 财务分析报告</title>
<style>
:root {{ --primary:#1a56db; --danger:#dc2626; --success:#16a34a; --bg:#f8fafc; --card:#fff; --border:#e2e8f0; --text:#1e293b; --text2:#64748b; }}
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif; background:var(--bg); color:var(--text); line-height:1.6; }}
.container {{ max-width:1400px; margin:0 auto; padding:20px; }}
.header {{ background:linear-gradient(135deg,#1e3a8a,#3b82f6); color:#fff; padding:40px; border-radius:16px; margin-bottom:24px; }}
.header h1 {{ font-size:28px; margin-bottom:8px; }}
.header .subtitle {{ opacity:0.85; font-size:14px; }}
.card {{ background:var(--card); border-radius:12px; padding:24px; margin-bottom:20px; box-shadow:0 1px 3px rgba(0,0,0,0.08); }}
.card h2 {{ font-size:18px; margin-bottom:16px; padding-bottom:8px; border-bottom:2px solid var(--primary); display:inline-block; }}
.card h3 {{ font-size:15px; color:var(--text2); margin:16px 0 8px; }}
.metrics-grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(200px,1fr)); gap:16px; margin-bottom:20px; }}
.metric-card {{ background:#f1f5f9; border-radius:10px; padding:16px; text-align:center; }}
.metric-card .label {{ font-size:12px; color:var(--text2); margin-bottom:4px; }}
.metric-card .value {{ font-size:24px; font-weight:700; }}
.metric-card .change {{ font-size:12px; margin-top:4px; }}
.positive {{ color:var(--success); }}
.negative {{ color:var(--danger); }}
table {{ width:100%; border-collapse:collapse; font-size:13px; overflow-x:auto; display:block; }}
th, td {{ padding:8px 10px; text-align:right; white-space:nowrap; border-bottom:1px solid var(--border); }}
th {{ background:#f1f5f9; font-weight:600; position:sticky; top:0; }}
td:first-child, th:first-child {{ text-align:left; font-weight:500; position:sticky; left:0; background:#fff; z-index:1; min-width:180px; }}
th:first-child {{ background:#f1f5f9; z-index:2; }}
tr:hover td {{ background:#f8fafc; }}
.forecast {{ background:#fffbeb !important; }}
.forecast-val {{ font-style:italic; color:#92400e; }}
.forecast-dot {{ color:#f59e0b; font-size:10px; vertical-align:super; cursor:help; }}
.analysis-box {{ background:#f8fafc; border-left:3px solid var(--primary); padding:14px 18px; border-radius:0 8px 8px 0; margin:16px 0; font-size:14px; line-height:1.8; color:#334155; }}
.summary-box {{ background:linear-gradient(135deg,#eff6ff,#dbeafe); border-left:4px solid var(--primary); padding:16px 20px; border-radius:0 8px 8px 0; margin:16px 0; }}
.legend {{ display:flex; gap:16px; align-items:center; font-size:12px; color:var(--text2); margin:8px 0 16px; flex-wrap:wrap; }}
.legend-item {{ display:flex; align-items:center; gap:4px; }}
.legend-swatch {{ width:20px; height:2px; display:inline-block; }}
.legend-swatch.solid {{ background:#16a34a; }}
.legend-swatch.dashed {{ background:repeating-linear-gradient(90deg,#f59e0b 0,#f59e0b 4px,transparent 4px,transparent 6px); }}
.two-col {{ display:grid; grid-template-columns:1fr 1fr; gap:20px; }}
.risk-list {{ padding-left:20px; line-height:2; }}
.risk-list li {{ margin-bottom:4px; }}
@media (max-width:768px) {{ .two-col {{ grid-template-columns:1fr; }} .container {{ padding:10px; }} }}
@media print {{ body {{ background:#fff; }} .card {{ box-shadow:none; border:1px solid #ddd; page-break-inside:avoid; }} }}
</style>
</head>
<body>
<div class="container">
<div class="header">
<h1>📊 {title}</h1>
<div class="subtitle">财务分析报告 | 单位：亿元(CNY)</div>
</div>

<div class="card">
<h2>📋 核心摘要（{dates[latest]}年实际数据）</h2>
<div class="metrics-grid">
<div class="metric-card">
<div class="label">{dates[latest]}年营收</div>
<div class="value">{fmt(revenue[latest])}</div>
<div class="change">同比 {rv_yoy}</div>
</div>
<div class="metric-card">
<div class="label">{dates[latest]}年归母净利润</div>
<div class="value">{fmt(net_profit_parent[latest])}</div>
</div>
<div class="metric-card">
<div class="label">{dates[latest]}年毛利率</div>
<div class="value">{fmt_pct(gross_margin[latest])}</div>
</div>
<div class="metric-card">
<div class="label">{dates[latest]}年ROE</div>
<div class="value">{fmt_pct(roe[latest])}</div>
</div>
</div>
<div class="legend">
<div class="legend-item"><div class="legend-swatch solid"></div> 实际值趋势</div>
<div class="legend-item"><div class="legend-swatch dashed"></div> 预测值趋势</div>
<div class="legend-item"><span class="forecast-val">斜体</span> 预测数值</div>
<div class="legend-item"><span class="forecast-dot">⟡</span> 预测标记</div>
</div>
</div>

<div class="card">
<h2>📈 一、盈利能力分析</h2>
<table>
{table_header()}
{table_row("营业总收入", revenue)}
{table_row("营业利润", op_profit)}
{table_row("归母净利润", net_profit_parent)}
{table_row("研发支出", rd)}
{table_row("EBITDA", ebitda)}
</table>
<h3 style="margin:16px 0 8px;">利润率指标</h3>
<table>
{table_header()}
{table_row("毛利率", gross_margin, True)}
{table_row("净利率", net_margin, True)}
{table_row("ROE(摊薄)", roe, True)}
{table_row("ROA", roa, True)}
</table>
<div class="analysis-box">{profit_text}</div>
</div>

<div class="card">
<h2>🏦 二、资产负债分析</h2>
<table>
{table_header()}
{table_row("总资产", total_assets)}
{table_row("流动资产", current_assets)}
{table_row("总负债", total_liab)}
{table_row("股东权益", equity)}
{table_row("资产负债率", debt_ratio, True)}
</table>
<div class="analysis-box">{bs_text}</div>
</div>

<div class="card">
<h2>💰 三、现金流分析</h2>
<table>
{table_header()}
{table_row("经营活动现金流", ocf)}
{table_row("投资活动现金流", icf)}
{table_row("筹资活动现金流", fcf_finance)}
{table_row("资本开支", capex)}
{table_row("自由现金流(OCF-CapEx)", fcf)}
{table_row("期末现金余额", cash_end)}
</table>
<div class="analysis-box">{cf_text}</div>
</div>

<div class="card">
<h2>📊 四、每股指标与效率</h2>
<table>
{table_header()}
{table_row("EPS(元)", eps)}
{table_row("每股净资产(元)", bps)}
{table_row("资产周转率(倍)", asset_turnover)}
{table_row("员工人数", employees)}
</table>
<div class="analysis-box">{ps_text}</div>
</div>

<div class="card">
<h2>🏭 五、行业分析</h2>
{industry_html}
</div>

<div class="card">
<h2>⚠️ 六、风险与机遇分析</h2>
{risk_html}
</div>

<div style="text-align:center;color:var(--text2);font-size:12px;padding:20px;">
📅 数据来源：公司财务报告 | 仅供参考，不构成投资建议
</div>
</div>
</body>
</html>"""
    return html


def html_to_pdf(html_path, pdf_path):
    for cmd in [
        ["wkhtmltopdf", "--quiet", "--enable-local-file-access", "--page-size", "A4", "--orientation", "Landscape", html_path, pdf_path],
    ]:
        try:
            subprocess.run(cmd, check=True, capture_output=True)
            return True
        except (FileNotFoundError, subprocess.CalledProcessError):
            pass
    for browser in ["chromium-browser", "chromium", "google-chrome"]:
        try:
            subprocess.run([browser, "--headless", "--disable-gpu", "--no-sandbox", f"--print-to-pdf={pdf_path}", html_path], check=True, capture_output=True)
            return True
        except (FileNotFoundError, subprocess.CalledProcessError):
            continue
    return False


def html_to_docx(html_path, docx_path):
    try:
        subprocess.run(["pandoc", html_path, "-o", docx_path, "--from=html"], check=True, capture_output=True)
        return True
    except (FileNotFoundError, subprocess.CalledProcessError):
        return False


def html_to_markdown(html_path, md_path):
    try:
        subprocess.run(["pandoc", html_path, "-o", md_path, "--from=html", "--to=gfm"], check=True, capture_output=True)
        return True
    except (FileNotFoundError, subprocess.CalledProcessError):
        pass
    try:
        os.system("pip install markdownify -q")
        from markdownify import markdownify
        with open(html_path) as f:
            md = markdownify(f.read(), heading_style="ATX")
        with open(md_path, "w") as f:
            f.write(md)
        return True
    except:
        return False


def main():
    parser = argparse.ArgumentParser(description="Generate financial analysis report")
    parser.add_argument("input", help="Input Excel file path")
    parser.add_argument("-o", "--output-format", default="pdf", choices=["pdf", "doc", "md", "html"])
    parser.add_argument("--output-dir", default=".", help="Output directory")
    parser.add_argument("--company", default="", help="Company name")
    parser.add_argument("--ticker", default="", help="Stock ticker")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    base_name = args.company.replace(" ", "_") if args.company else os.path.splitext(os.path.basename(args.input))[0]

    rows, header_row = parse_excel(args.input)
    indices, report_types = identify_periods(rows)
    html = generate_html(rows, header_row, indices, report_types, args.company, args.ticker)

    html_path = os.path.join(args.output_dir, f"{base_name}_report.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"✅ HTML report: {html_path} ({len(html):,} bytes)")

    if args.output_format != "html":
        ext_map = {"pdf": "pdf", "doc": "docx", "md": "md"}
        converter = {"pdf": html_to_pdf, "doc": html_to_docx, "md": html_to_markdown}.get(args.output_format)
        if converter:
            ext = ext_map[args.output_format]
            out_path = os.path.join(args.output_dir, f"{base_name}_report.{ext}")
            if converter(html_path, out_path):
                print(f"✅ {ext.upper()} report: {out_path}")
            else:
                print(f"⚠️  {ext.upper()} conversion failed. HTML report available at: {html_path}")


if __name__ == "__main__":
    main()
